# #Task1

a=b'r\xc3\xa9sum\xc3\xa9'
b = a.decode("utf-8")
print(b)
c = b.encode("Latin1")
print(c)
f=c.decode("Latin1")
print(f)

# #Task2

a = str(input("Введи строку 1:"))
b = str(input("Введи строку 2:"))
c = str(input("Введи строку 3:"))
d = str(input("Введи строку 4:"))

with open("str.py","w") as new_file:
    new_file.write(a + "\n" + b + "\n")
with open("str.py","a") as new_file:
    new_file.write(c + "\n" + d + "\n")

# #Task3

import json
dict = {
    121212 : ("Serg",40),
    232323 : ("Olya",37),
    343434 : ("Gleb",10),
    454545 : ("Zhorik",17),
    565656 : ("Mara",6)
}

with open("dict.json","w") as new_file:
    json.dump(dict,new_file,indent= 4)

#Task4

import json
import csv

phones = {
    "Serg" : 375298084411,
    "Olya" : 375294877133,
    "Gleb" : 375255118578,
    "Zhorik" : 375297474720,
    "Mara": 375292183345
}
with open("dict.json","r") as file:
    read_data = json.load(file)

with open("dict.csv", "w", newline="") as new_file:
        main_names = ["id", "name", "age", "phone"]
        writer = csv.DictWriter(new_file, fieldnames = main_names)
        writer.writeheader()
        for id, (name, age) in read_data.items():
            phone = phones.get(name, "")
            writer.writerow({'id': id, 'name': name, 'age': age, 'phone': phone})

#Task5

import csv
from openpyxl import Workbook

list_data = []
with open("dict.csv","r") as new_file:
    reader = csv.reader(new_file)
    headers = next(reader)
    for row in reader:
        list_data.append(row[:2] + row[3:])

wb = Workbook()
ws = wb.active

for row_index,value in enumerate(headers[:3],start = 2):
    ws.cell(row = row_index, column = 1 , value = value)
    ws.cell(row = 4, column = 1, value = headers[3])

for col_index,item in enumerate(range(1,7),start = 2):
    ws.cell(row = 1, column = col_index, value = f"Person{item}" )
for col_index, item in enumerate(list_data, start=2):
    for row_index, value in enumerate(item, start=2):
        ws.cell(row=row_index, column=col_index, value=value)

wb.save("dict.xlsx")